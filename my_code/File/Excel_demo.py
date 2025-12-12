import openpyxl
import datetime

# 加载工作簿
wb = openpyxl.load_workbook(r'my_code\File\res\2022年股票数据.xlsx')
print(wb.sheetnames)
sheet = wb.worksheets[0]
print(sheet.dimensions)
print(sheet.max_row, sheet.max_column)

# 获取指定单元格的值
print(sheet.cell(3,3).value)
print(sheet['C3'].value)
print(sheet['G255'].value)

# 获取多个单元格
print(sheet['A2:C5'])

